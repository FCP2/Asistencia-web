# app.py
import os
from datetime import datetime, date, time as dtime, timedelta
from flask import Flask, render_template, request, jsonify
from sqlalchemy.orm import Session
from sqlalchemy import select, func
from db import (
    engine, SessionLocal, Persona, Invitacion, Notificacion
)
import os, uuid, mimetypes
from werkzeug.utils import secure_filename
from flask import send_from_directory
from io import BytesIO
from flask import send_file
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import re
from typing import Optional
# -----------------------------------------------------------------------------
# Helpers de formato
# ----------------------------------------------------------------------------

UPLOAD_FOLDER = os.getenv("UPLOAD_FOLDER", os.path.join(os.path.dirname(__file__), "uploads"))
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
ALLOWED_EXTS = {"pdf", "jpg", "jpeg", "png"}

def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTS

# ---------- PARSEOS (1 sola fuente de verdad) ----------

def parse_date_iso(s: Optional[str]) -> Optional[date]:
    """Estricto: solo ISO YYYY-MM-DD."""
    if not s:
        return None
    try:
        return date.fromisoformat(s.strip())
    except ValueError:
        return None

def parse_date_flexible(s: Optional[str]) -> Optional[date]:
    """Tolerante: ISO (YYYY-MM-DD) o dd/mm/aaaa o dd/mm/aa."""
    if not s:
        return None
    s = s.strip()
    # 1) ISO
    d = parse_date_iso(s)
    if d:
        return d
    # 2) dd/mm/aaaa o dd/mm/aa
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

def parse_time_flexible(s: Optional[str]) -> Optional[dtime]:
    """
    Acepta:
      - HH:MM
      - HH:MM:SS
      - 'h:mm am/pm' / 'hh am' (ej. '3 pm', '3:05 PM')
    """
    if not s:
        return None
    v = s.strip().lower()

    # 12h con am/pm (permite '3 pm', '03:05 pm', '12:30am', etc.)
    m = re.match(r"^(\d{1,2})(?::(\d{1,2}))?(?::(\d{1,2}))?\s*(am|pm)$", v)
    if m:
        hh = int(m.group(1))
        mm = int(m.group(2) or 0)
        ss = int(m.group(3) or 0)
        ampm = m.group(4)
        if hh == 12:
            hh = 0
        if ampm == "pm":
            hh += 12
        return dtime(hh, mm, ss)

    # 24h HH:MM(:SS)
    parts = v.split(":")
    try:
        if len(parts) == 2:
            return dtime(int(parts[0]), int(parts[1]))
        if len(parts) >= 3:
            return dtime(int(parts[0]), int(parts[1]), int(parts[2]))
    except ValueError:
        return None
    return None

# ---------- FORMATOS (puedes dejarlos igual) ----------

def fmt_date(d: Optional[date]) -> str:
    return d.strftime("%d/%m/%y") if d else ""

def fmt_time(t: Optional[dtime]) -> str:
    return t.strftime("%H:%M") if t else ""

def fmt_dt(dtobj: Optional[datetime]) -> str:
    return dtobj.strftime("%d/%m/%y %H:%M") if dtobj else ""

# ---------- Utilidad ----------

def minutes_diff(t1: Optional[dtime], t2: Optional[dtime]) -> int:
    """Diferencia absoluta en minutos entre dos times; 'infinito' si falta alguno."""
    if not (t1 and t2):
        return 10**6
    dt1 = datetime.combine(date(2000, 1, 1), t1)
    dt2 = datetime.combine(date(2000, 1, 1), t2)
    return int(abs((dt1 - dt2).total_seconds()) // 60)
# -----------------------------------------------------------------------------
# Serializador de invitaciones
# -----------------------------------------------------------------------------
def inv_to_dict(inv: Invitacion) -> dict:
    # días hacia el evento (para highlight en UI)
    dias = (inv.fecha - date.today()).days if inv.fecha else None
    # nombre vivo desde la relación (si existe), si no, cae al texto desnormalizado:
    asignado_nombre = (inv.persona.nombre if getattr(inv, "persona", None) and inv.persona else None) \
                      or (inv.asignado_a or "")

    return {
        "ID": inv.id,
        "PersonaID": inv.persona_id,
        "Evento": inv.evento or "",
        "Convoca Cargo": inv.convoca_cargo or "",
        "Convoca": inv.convoca or "",
        "Partido Político": inv.partido_politico or "",
        # Inputs/queries
        "Fecha": inv.fecha.isoformat() if inv.fecha else None,
        "Hora": inv.hora.strftime("%H:%M") if inv.hora else None,
        # Alias explícitos (por si en front usas getFecha/getHora)
        "FechaISO": inv.fecha.isoformat() if inv.fecha else None,
        "HoraISO": inv.hora.strftime("%H:%M") if inv.hora else None,
        # UI formateada
        "FechaFmt": fmt_date(inv.fecha),
        "HoraFmt": fmt_time(inv.hora),
        "Municipio/Dependencia": inv.municipio or "",
        "Lugar": inv.lugar or "",
        "Estatus": inv.estatus or "Pendiente",
        "Asignado A": asignado_nombre,         # <— AQUÍ
        "PersonaNombre": (inv.persona.nombre if inv.persona else None),  # opcional, por claridad        "Rol": inv.rol or "",
        "Observaciones": inv.observaciones or "",
        "Fecha Asignación": fmt_dt(inv.fecha_asignacion),
        "Última Modificación": fmt_dt(inv.ultima_modificacion),
        "Modificado Por": inv.modificado_por or "",
        # Archivo
        "ArchivoURL": inv.archivo_url or "",
        "ArchivoNombre": inv.archivo_nombre or "",
        "ArchivoMime": inv.archivo_mime or "",
        "ArchivoTamano": inv.archivo_tamano or 0,
        "ArchivoTS": fmt_dt(inv.archivo_ts),
        # Nuevo: días hacia el evento para el highlight
        "DiasParaEvento": dias,
    }
    


# -----------------------------------------------------------------------------
# Notificaciones: snapshot en tabla notificaciones
# -----------------------------------------------------------------------------
def add_notif(db: Session, inv: Invitacion, campo: str, old_val: str | None, new_val: str | None, comentario: str = ""):
    n = Notificacion(
        ts = datetime.now(),
        invitacion_id = str(inv.id),

        # snapshot textual principal
        evento = inv.evento or "",
        convoca = inv.convoca or "",
        estatus = inv.estatus or "",
        asignado_a_nombre = inv.asignado_a or "",
        rol = inv.rol or "",
        campo = campo or "",
        valor_anterior = old_val or "",
        valor_nuevo = new_val or "",
        comentario = comentario or "",

        # snapshot de datos del evento
        fecha = inv.fecha,
        hora = inv.hora,
        municipio = inv.municipio or "",
        lugar = inv.lugar or "",
        convoca_cargo=inv.convoca_cargo,
        enviado = False,
        enviado_ts = None
    )
    db.add(n)

# -----------------------------------------------------------------------------
# Flask app
# -----------------------------------------------------------------------------
app = Flask(__name__)

@app.get("/")
def home():
    return render_template("index.html")

@app.get("/health")
def health():
    return {"ok": True, "ts": datetime.utcnow().isoformat()}

# -----------------------------------------------------------------------------
# PERSONAS (Catálogo)
# -----------------------------------------------------------------------------
@app.get("/api/catalog")
def api_catalog():
    """Lista personas activas para el combo de selección."""
    db = SessionLocal()
    try:
        personas = (db.query(Persona)
                      .filter(Persona.activo == True)
                      .order_by(Persona.nombre.asc())
                      .all())
        rows = [{
            "ID": p.id,
            "Nombre": p.nombre or "",
            "Cargo": p.cargo or "",
            "Teléfono": p.telefono or "",
            "Correo": p.correo or "",
            "Unidad/Región": p.unidad_region or "",
        } for p in personas]
        return jsonify(rows)
    finally:
        db.close()

@app.post("/api/person/create")
def api_person_create():
    data = request.get_json() or {}
    nombre = (data.get("Nombre") or "").strip()
    cargo  = (data.get("Cargo") or "").strip()

    if not nombre or not cargo:
        return jsonify({"ok": False, "error": "Nombre y Cargo son obligatorios"}), 400

    db = SessionLocal()
    try:
        p = Persona(
            nombre=nombre,
            cargo=cargo,
            telefono=(data.get("Teléfono") or "").strip(),
            correo=(data.get("Correo") or "").strip(),
            unidad_region=(data.get("Unidad/Región") or "").strip(),
            activo=True
        )
        db.add(p)
        db.commit()
        return jsonify({"ok": True, "id": p.id})
    except Exception as e:
        db.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500
    finally:
        db.close()

@app.post("/api/person/update")
def api_person_update():
    data = request.get_json() or {}
    pid = data.get("ID")
    if not pid:
        return jsonify({"ok": False, "error": "Falta ID"}), 400

    db = SessionLocal()
    try:
        p = db.get(Persona, int(pid))
        if not p:
            return jsonify({"ok": False, "error": "Persona no encontrada"}), 404

        for key, attr in [
            ("Nombre", "nombre"),
            ("Cargo", "cargo"),
            ("Teléfono", "telefono"),
            ("Correo", "correo"),
            ("Unidad/Región", "unidad_region"),
        ]:
            val = data.get(key)
            if val is not None:
                setattr(p, attr, val.strip() if isinstance(val, str) else val)

        db.commit()
        return jsonify({"ok": True, "persona": {"ID": p.id, "Nombre": p.nombre}})
    except Exception as e:
        db.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500
    finally:
        db.close()

@app.post("/api/person/delete")
def api_person_delete():
    """Elimina persona y pone en Pendiente sus invitaciones, conservando el nombre en 'Asignado A'."""
    data = request.get_json() or {}
    pid = data.get("ID")
    if not pid:
        return jsonify({"ok": False, "error": "Falta ID"}), 400

    db = SessionLocal()
    try:
        p = db.get(Persona, int(pid))
        if not p:
            return jsonify({"ok": False, "error": "Persona no encontrada"}), 404

        # 1) Desasignar invitaciones de esta persona.
        #    Dejamos 'Asignado A' como está (nombre permanece), solo se quita persona_id y se pone Pendiente.
        invs = (
            db.query(Invitacion)
              .filter(Invitacion.persona_id == p.id)
              # si quieres excluir canceladas, descomenta:
              # .filter(Invitacion.estatus != "Cancelado")
              .all()
        )

        for inv in invs:
            prev_estatus = inv.estatus

            inv.persona_id = None                     # se rompe el vínculo
            inv.estatus = "Pendiente"                 # vuelve a pendiente
            # mantenemos inv.asignado_a (nombre) y inv.rol tal cual
            inv.observaciones = ((inv.observaciones or "") +
                                 (" | " if inv.observaciones else "") +
                                 "Auto-desasignación: persona eliminada")
            inv.ultima_modificacion = datetime.now()
            inv.modificado_por = "ATIapp"

            # notificación para que el bot la detecte (cambio de estatus)
            add_notif(
                db, inv,
                campo="Estatus",
                old_val=prev_estatus or "",
                new_val="Pendiente",
                comentario=f"Persona eliminada: {p.nombre or ''}"
            )

        # 2) Ahora sí eliminamos la persona
        db.delete(p)
        db.commit()
        return jsonify({"ok": True, "invitaciones_actualizadas": len(invs)})

    except Exception as e:
        db.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500
    finally:
        db.close()

# -----------------------------------------------------------------------------
# INVITACIONES
# -----------------------------------------------------------------------------
@app.get("/api/invitations")
def api_invitations():
    """Lista invitaciones. Soporta ?status=... y ?date_from=YYYY-MM-DD|dd/mm/aaaa & ?date_to=..."""
    status    = (request.args.get("status") or "").strip()
    date_from = parse_date_flexible(request.args.get("date_from"))
    date_to   = parse_date_flexible(request.args.get("date_to"))

    # Si el usuario invirtió el rango, lo corregimos
    if date_from and date_to and date_from > date_to:
        date_from, date_to = date_to, date_from

    db = SessionLocal()
    try:
        q = db.query(Invitacion)
        if status:
            q = q.filter(Invitacion.estatus == status)
        if date_from:
            q = q.filter(Invitacion.fecha >= date_from)
        if date_to:
            q = q.filter(Invitacion.fecha <= date_to)

        invs = (q.order_by(Invitacion.fecha.desc().nullslast(),
                           Invitacion.hora.desc().nullslast(),
                           Invitacion.id.desc())
                  .all())
        return jsonify([inv_to_dict(i) for i in invs])
    finally:
        db.close()

@app.get("/api/invitation/<int:inv_id>")
def api_inv_get(inv_id: int):
    """Detalle de una invitación."""
    db = SessionLocal()
    try:
        inv = db.get(Invitacion, inv_id)
        if not inv:
            return jsonify({"ok": False, "error": "Invitación no encontrada"}), 404
        return jsonify(inv_to_dict(inv))
    finally:
        db.close()

@app.post("/api/invitation/create")
def api_invitation_create():
    # Campos desde FormData
    f_fecha   = request.form.get("fecha")
    f_hora    = request.form.get("hora")
    f_evento  = request.form.get("evento")
    f_ccargo  = request.form.get("convoca_cargo")  # "Diputado(a)" etc.
    f_convoca = request.form.get("convoca")
    f_partido = request.form.get("partido_politico")
    f_muni    = request.form.get("municipio")
    f_lugar   = request.form.get("lugar")
    f_obs     = request.form.get("observaciones")

    if not (f_fecha and f_hora and f_evento and f_ccargo and f_convoca and f_muni and f_lugar):
        return jsonify({"ok": False, "error": "Faltan campos obligatorios"}), 400

    archivo = request.files.get("archivo")
    archivo_url = archivo_nombre = archivo_mime = None
    archivo_tamano = None
    archivo_ts = None

    if archivo and archivo.filename and allowed_file(archivo.filename):
        ext = archivo.filename.rsplit(".", 1)[1].lower()
        safe_orig = secure_filename(archivo.filename)
        unique = f"{uuid.uuid4().hex}.{ext}"
        save_path = os.path.join(UPLOAD_FOLDER, unique)
        archivo.save(save_path)

        archivo_url = f"/uploads/{unique}"
        archivo_nombre = safe_orig
        archivo_mime = mimetypes.guess_type(safe_orig)[0] or archivo.mimetype or "application/octet-stream"
        archivo_tamano = os.path.getsize(save_path)
        archivo_ts = datetime.utcnow()

    inv = Invitacion(
    fecha=parse_date_flexible(f_fecha),
    hora=parse_time_flexible(f_hora),
    evento=f_evento,
    convoca_cargo=f_ccargo,
    convoca=f_convoca,
    partido_politico=f_partido,
    municipio=f_muni,
    lugar=f_lugar,
    observaciones=f_obs or "",
    archivo_url=archivo_url,
    archivo_nombre=archivo_nombre,
    archivo_mime=archivo_mime,
    archivo_tamano=archivo_tamano,
    archivo_ts=archivo_ts,
    ultima_modificacion=datetime.utcnow(),
    modificado_por="atiapp",
)


    db = SessionLocal()
    try:
        db.add(inv)
        db.commit()
        return jsonify({"ok": True, "id": inv.id})
    except Exception as e:
        db.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500
    finally:
        db.close()

@app.post("/api/invitation/update")
def api_invitation_update():
    inv_id = request.form.get("id") or request.form.get("ID")
    if not inv_id:
        return jsonify({"ok": False, "error": "Falta ID"}), 400

    db = SessionLocal()
    try:
        inv = db.get(Invitacion, int(inv_id))
        if not inv:
            return jsonify({"ok": False, "error": "Invitación no encontrada"}), 404

        # Campos (permitimos que falten)
        f_fecha   = request.form.get("fecha") or request.form.get("Fecha")
        f_hora    = request.form.get("hora")  or request.form.get("Hora")
        f_evento  = request.form.get("evento") or request.form.get("Evento")
        f_ccargo  = request.form.get("convoca_cargo") or request.form.get("Convoca Cargo")
        f_convoca = request.form.get("convoca") or request.form.get("Convoca")
        f_partido = request.form.get("partido_politico") or request.form.get("Partido Político")
        f_muni    = request.form.get("municipio") or request.form.get("Municipio/Dependencia")
        f_lugar   = request.form.get("lugar") or request.form.get("Lugar")
        f_obs     = request.form.get("observaciones") or request.form.get("Observaciones")
        eliminar_archivo = (request.form.get("eliminar_archivo") or "").lower() == "true"

        if f_evento is not None:  inv.evento = f_evento
        if f_ccargo is not None:  inv.convoca_cargo = f_ccargo
        if f_convoca is not None: inv.convoca = f_convoca
        if f_partido is not None: inv.partido_politico = f_partido
        if f_muni is not None:    inv.municipio = f_muni
        if f_lugar is not None:   inv.lugar = f_lugar
        if f_obs is not None:     inv.observaciones = f_obs
        if f_fecha is not None:   inv.fecha = parse_date_flexible(f_fecha)
        if f_hora is not None:    inv.hora  = parse_time_flexible(f_hora)
        file = request.files.get("archivo")
        if eliminar_archivo:
            inv.archivo_url = inv.archivo_nombre = inv.archivo_mime = None
            inv.archivo_tamano = None
            inv.archivo_ts = None

        if file and file.filename and allowed_file(file.filename):
            ext = file.filename.rsplit(".", 1)[1].lower()
            safe_orig = secure_filename(file.filename)
            unique = f"{uuid.uuid4().hex}.{ext}"
            save_path = os.path.join(UPLOAD_FOLDER, unique)
            file.save(save_path)

            inv.archivo_url = f"/uploads/{unique}"
            inv.archivo_nombre = safe_orig
            inv.archivo_mime = mimetypes.guess_type(safe_orig)[0] or file.mimetype or "application/octet-stream"
            inv.archivo_tamano = os.path.getsize(save_path)
            inv.archivo_ts = datetime.utcnow()

        inv.ultima_modificacion = datetime.utcnow()
        inv.modificado_por = "atiapp"

        db.commit()
        return jsonify({"ok": True})
    except Exception as e:
        db.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500
    finally:
        db.close()

@app.post("/api/invitation/delete")
def api_inv_delete():
    """Eliminar invitación."""
    data = request.get_json() or {}
    inv_id = data.get("ID") or data.get("id")
    if not inv_id:
        return jsonify({"ok": False, "error": "Falta ID"}), 400

    db = SessionLocal()
    try:
        inv = db.get(Invitacion, int(inv_id))
        if not inv:
            return jsonify({"ok": False, "error": "Invitación no encontrada"}), 404
        db.delete(inv)
        db.commit()
        return jsonify({"ok": True})
    except Exception as e:
        db.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500
    finally:
        db.close()
        
@app.get("/api/stats")
def api_stats():
    date_from = parse_date_flexible(request.args.get("date_from"))
    date_to   = parse_date_flexible(request.args.get("date_to"))
    if date_from and date_to and date_from > date_to:
        date_from, date_to = date_to, date_from

    db = SessionLocal()
    try:
        filters = []
        if date_from:
            filters.append(Invitacion.fecha >= date_from)
        if date_to:
            filters.append(Invitacion.fecha <= date_to)

        rows = (db.query(Invitacion.estatus, func.count(Invitacion.id))
                  .filter(*filters)
                  .group_by(Invitacion.estatus)
                  .all())

        counts = {"Pendiente": 0, "Confirmado": 0, "Sustituido": 0, "Cancelado": 0}
        for est, c in rows:
            counts[est or "Pendiente"] = c
        return jsonify(counts)
    finally:
        db.close()
# -----------------------------------------------------------------------------
# ACCIONES: asignar / sustituir / status / cancelar
# -----------------------------------------------------------------------------
@app.post("/api/assign")
def api_assign():
    """Confirmar y asignar persona por persona_id (con chequeo de conflictos)."""
    data = request.get_json() or {}
    inv_id_raw     = data.get("id")
    persona_id_raw = data.get("persona_id")
    rol_in     = (data.get("rol") or "").strip()
    comentario = (data.get("comentario") or "").strip()
    force      = bool(data.get("force", False))

    # Cast seguro de IDs
    try:
        inv_id = int(inv_id_raw)
        persona_id = int(persona_id_raw)
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "id/persona_id inválidos"}), 400

    db = SessionLocal()
    try:
        inv = db.get(Invitacion, inv_id)
        p   = db.get(Persona, persona_id)
        if not inv or not p:
            return jsonify({"ok": False, "error": "Invitación o persona no encontrada"}), 404

        # === Chequeo de conflicto (si hay fecha/hora) salvo force ===
        if inv.fecha and inv.hora and not force:
            q = (db.query(Invitacion)
                    .filter(Invitacion.persona_id == p.id)
                    .filter(Invitacion.fecha == inv.fecha)
                    .filter(Invitacion.estatus.in_(["Confirmado","Sustituido"]))
                    .filter(Invitacion.id != inv.id))
            matches = q.all()

            level = "none"
            conflicts = []
            order = {"none":0, "tight2h":1, "tight1h":2, "hard":3}

            for m in matches:
                # si alguna hora es None, la tratamos como no comparable (sin conflicto)
                dm = minutes_diff(m.hora, inv.hora) if (m.hora and inv.hora) else 10**6
                if dm == 0:
                    lev = "hard"
                elif dm <= 60:
                    lev = "tight1h"
                elif dm <= 60:
                    lev = "tight2h"
                else:
                    lev = "none"

                if lev != "none":
                    conflicts.append({
                        "ID": m.id,
                        "Evento": m.evento or "",
                        "FechaFmt": fmt_date(m.fecha),
                        "HoraFmt": fmt_time(m.hora),
                        "Estatus": m.estatus or "",
                        "Lugar": m.lugar or ""
                    })
                    if order[lev] > order[level]:
                        level = lev

            if level != "none":
                return jsonify({
                    "ok": False, "conflict": True,
                    "level": level, "conflicts": conflicts
                }), 409

        # === Aplicar asignación ===
        prev_estatus = inv.estatus
        prev_asig    = inv.asignado_a
        prev_rol     = inv.rol

        inv.persona_id = p.id
        inv.asignado_a = p.nombre
        inv.rol        = (rol_in if rol_in else (p.cargo or ""))
        inv.estatus    = "Confirmado"
        if comentario:
            inv.observaciones = ((inv.observaciones or "")
                                 + (" | " if inv.observaciones else "")
                                 + comentario)
        inv.fecha_asignacion    = datetime.now()
        inv.ultima_modificacion = datetime.now()
        inv.modificado_por      = "atiapp"

        add_notif(db, inv, "Asignado A", prev_asig or "", inv.asignado_a or "", comentario)
        if prev_rol != inv.rol:
            add_notif(db, inv, "Rol", prev_rol or "", inv.rol or "", comentario)
        if prev_estatus != inv.estatus:
            add_notif(db, inv, "Estatus", prev_estatus or "", inv.estatus or "", comentario)

        db.commit()
        return jsonify({"ok": True})

    except Exception as e:
        db.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500
    finally:
        db.close()

@app.post("/api/reassign")
def api_reassign():
    """Sustituir (cambia persona y estatus a Sustituido)."""
    data = request.get_json() or {}
    inv_id = data.get("id")
    persona_id = data.get("persona_id")
    rol_in = (data.get("rol") or "").strip()
    comentario = (data.get("comentario") or "Sustitución por instrucción").strip()
    force = bool(data.get("force", False))

    if not inv_id or not persona_id:
        return jsonify({"ok": False, "error": "Faltan campos: id, persona_id"}), 400

    db = SessionLocal()
    try:
        inv = db.get(Invitacion, int(inv_id))
        p = db.get(Persona, int(persona_id))
        if not inv or not p:
            return jsonify({"ok": False, "error": "Invitación o persona no encontrada"}), 404
        
        # === Chequeo de conflicto, salvo que venga force ===
        if inv.fecha and inv.hora and not force:
            # Llamamos internamente la lógica del endpoint
            q = (db.query(Invitacion)
                    .filter(Invitacion.persona_id == p.id)
                    .filter(Invitacion.fecha == inv.fecha)
                    .filter(Invitacion.estatus.in_(["Confirmado","Sustituido"]))
                    .filter(Invitacion.id != inv.id))
            matches = q.all()

            level = "none"; conflicts=[]
            for m in matches:
                dm = minutes_diff(m.hora, inv.hora)
                if dm == 0: lev="hard"
                elif dm <= 60: lev="tight1h"
                elif dm <= 60: lev="tight2h"
                else: lev="none"
                if lev!="none":
                    conflicts.append({
                        "ID": m.id, "Evento": m.evento or "",
                        "FechaFmt": fmt_date(m.fecha), "HoraFmt": fmt_time(m.hora),
                        "Estatus": m.estatus or "", "Lugar": m.lugar or ""
                    })
                    order={"none":0,"tight2h":1,"tight1h":2,"hard":3}
                    if order[lev]>order[level]: level=lev

            if level != "none":
                return jsonify({"ok": False, "conflict": True, "level": level, "conflicts": conflicts}), 409

        prev_estatus = inv.estatus
        prev_asig = inv.asignado_a
        prev_rol = inv.rol

        inv.persona_id = p.id
        inv.asignado_a = p.nombre
        inv.rol = (rol_in if rol_in else (p.cargo or ""))
        inv.estatus = "Sustituido"
        if comentario:
            inv.observaciones = ((inv.observaciones or "") + (" | " if inv.observaciones else "") + comentario)
        inv.ultima_modificacion = datetime.now()
        inv.modificado_por = "atiapp"

        add_notif(db, inv, "Asignado A", prev_asig or "", inv.asignado_a or "", comentario)
        if prev_rol != inv.rol:
            add_notif(db, inv, "Rol", prev_rol or "", inv.rol or "", comentario)
        if prev_estatus != inv.estatus:
            add_notif(db, inv, "Estatus", prev_estatus or "", inv.estatus or "", comentario)

        db.commit()
        return jsonify({"ok": True})
    except Exception as e:
        db.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500
    finally:
        db.close()

@app.post("/api/status")
def api_status():
    data = request.get_json() or {}
    inv_id = data.get("id")
    nuevo = (data.get("estatus") or "").strip()  # Pendiente, Confirmado, Sustituido, Cancelado
    comentario = (data.get("comentario") or "Cambio de estatus").strip()
    if not (inv_id and nuevo):
        return jsonify({"ok": False, "error": "Faltan campos"}), 400

    db = SessionLocal()
    try:
        inv = db.get(Invitacion, int(inv_id))
        if not inv:
            return jsonify({"ok": False, "error": "ID no encontrado"}), 404

        prev_estatus = inv.estatus
        prev_asig    = inv.asignado_a or ""
        prev_rol     = inv.rol or ""

        inv.estatus = nuevo
        inv.ultima_modificacion = datetime.utcnow()
        inv.modificado_por = "webapp"

        # Si regresamos a Pendiente, limpiamos los campos de asignación
        if nuevo == "Pendiente":
            inv.persona_id = None
            inv.asignado_a = ""
            inv.rol = ""
            inv.fecha_asignacion = None
            # opcional: también limpiar observaciones si quieres:
            # inv.observaciones = ""

        # Agrega comentario a observaciones (opcional)
        if comentario:
            inv.observaciones = ((inv.observaciones or "") + (" | " if inv.observaciones else "") + comentario)

        # Notificaciones (snapshots)
        add_notif(db, inv, "Estatus", prev_estatus or "", inv.estatus or "", comentario)
        if nuevo == "Pendiente":
            if prev_asig:
                add_notif(db, inv, "Asignado A", prev_asig, "", "Se limpió la asignación")
            if prev_rol:
                add_notif(db, inv, "Rol", prev_rol, "", "Se limpió la asignación")

        db.commit()
        return jsonify({"ok": True})
    except Exception as e:
        db.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500
    finally:
        db.close()

@app.post("/api/cancel")
def api_cancel():
    """Atajo para cancelar con motivo."""
    data = request.get_json() or {}
    inv_id = data.get("id")
    motivo = (data.get("comentario") or "Cancelado por indicación").strip()

    if not inv_id:
        return jsonify({"ok": False, "error": "Falta id"}), 400

    db = SessionLocal()
    try:
        inv = db.get(Invitacion, int(inv_id))
        if not inv:
            return jsonify({"ok": False, "error": "Invitación no encontrada"}), 404

        prev_estatus = inv.estatus
        prev_obs = inv.observaciones or ""

        inv.estatus = "Cancelado"
        inv.observaciones = (prev_obs + (" | " if prev_obs else "") + f"Motivo cancelación: {motivo}")
        inv.ultima_modificacion = datetime.now()
        inv.modificado_por = "atiapp"

        add_notif(db, inv, "Estatus", prev_estatus or "", "Cancelado", motivo)

        db.commit()
        return jsonify({"ok": True})
    except Exception as e:
        db.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500
    finally:
        db.close()
        
@app.post("/api/check-conflict")
def api_check_conflict():
    """
    Checa si persona_id tiene otra invitación Confirmada/Sustituida el mismo día.
    Body JSON: { persona_id, fecha (YYYY-MM-DD), hora (HH:MM), exclude_id? }
    Devuelve:
      { level: 'none'|'hard'|'tight1h'|'tight2h',
        conflicts: [ {ID, Evento, FechaFmt, HoraFmt, Estatus, Lugar} ] }
    """
    data = request.get_json() or {}
    persona_id = data.get("persona_id", type=int)
    fecha_str  = data.get("fecha") or data.get("Fecha") or data.get("FechaISO")
    hora_str   = data.get("hora")  or data.get("Hora")  or data.get("HoraISO")
    exclude_id = data.get("exclude_id", type=int)

    if not (persona_id and fecha_str and hora_str):
        return jsonify({"ok": False, "error": "Faltan persona_id/fecha/hora"}), 400

    try:
        fecha = date.fromisoformat(fecha_str)
        hh, mm = hora_str.split(":")[0:2]
        hora  = dtime(int(hh), int(mm))
    except Exception:
        return jsonify({"ok": False, "error": "Formato inválido de fecha/hora"}), 400

    db = SessionLocal()
    try:
        q = (db.query(Invitacion)
                .filter(Invitacion.persona_id == persona_id)
                .filter(Invitacion.fecha == fecha)
                .filter(Invitacion.estatus.in_(["Confirmado","Sustituido"])))
        if exclude_id:
            q = q.filter(Invitacion.id != exclude_id)

        matches = q.all()

        conflicts = []
        level = "none"
        for inv in matches:
            dm = minutes_diff(inv.hora, hora)
            # Clasificamos la severidad
            if dm == 0:
                lev = "hard"
            elif dm <= 60:
                lev = "tight1h"
            elif dm <= 60:
                lev = "tight2h"
            else:
                lev = "none"

            if lev != "none":
                conflicts.append({
                    "ID": inv.id,
                    "Evento": inv.evento or "",
                    "FechaFmt": fmt_date(inv.fecha),
                    "HoraFmt": fmt_time(inv.hora),
                    "Estatus": inv.estatus or "",
                    "Lugar": inv.lugar or ""
                })
                # Escala el nivel global (hard > tight1h > tight2h)
                order = {"none":0,"tight2h":1,"tight1h":2,"hard":3}
                if order[lev] > order[level]:
                    level = lev

        return jsonify({"ok": True, "level": level, "conflicts": conflicts})
    finally:
        db.close()
        
@app.get("/uploads/<path:fname>")
def uploads_serve(fname):
    return send_from_directory(UPLOAD_FOLDER, fname, as_attachment=False)

@app.get("/api/report/confirmados.xlsx")
def report_confirmados_xlsx():
    """Descarga un XLSX con las invitaciones Confirmadas."""
    db = SessionLocal()
    try:
        rows = (
            db.query(Invitacion)
              .filter(Invitacion.estatus == "Confirmado")
              .order_by(Invitacion.fecha.asc(), Invitacion.hora.asc())
              .all()
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Confirmados"

        headers = [
            "Municipio/Dependencia",
            "Partido Político",
            "Quien Convoca",
            "Asignado",
            "Cargo del Asignado",
            "Fecha",
            "Lugar",
            "Hora",
            "Convoca Cargo",
        ]
        ws.append(headers)

        def _fmt_date(d): return d.strftime("%d/%m/%y") if d else ""
        def _fmt_time(t): return t.strftime("%H:%M") if t else ""

        for inv in rows:
            ws.append([
                inv.municipio or "",
                inv.partido_politico or "",
                inv.convoca or "",
                inv.asignado_a or "",
                inv.rol or "",
                _fmt_date(inv.fecha),
                inv.lugar or "",
                _fmt_time(inv.hora),
                inv.convoca_cargo or "",
            ])

        # Encabezado bonito, autofiltro, freeze panes y ancho de columnas
        bold = Font(bold=True)
        fill = PatternFill("solid", fgColor="E9ECEF")
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = bold
            cell.fill = fill
            ws.column_dimensions[get_column_letter(c)].width = 24

        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
        ws.freeze_panes = "A2"

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        fname = f"reporte_confirmados_{datetime.utcnow().date().isoformat()}.xlsx"
        return send_file(
            bio,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=fname,
        )
    finally:
        db.close()
# -----------------------------------------------------------------------------
# CONTADORES (para dashboard)
# -----------------------------------------------------------------------------
@app.get("/api/counters")
def api_counters():
    """Regresa conteos por estatus para pintar los KPIs del header."""
    db = SessionLocal()
    try:
        rows = (db.query(Invitacion.estatus, func.count(Invitacion.id))
                  .group_by(Invitacion.estatus)
                  .all())
        counts = {"Pendiente": 0, "Confirmado": 0, "Sustituido": 0, "Cancelado": 0}
        for est, c in rows:
            counts[est or "Pendiente"] = c
        counts["Total"] = sum(counts.values())
        return jsonify(counts)
    finally:
        db.close()

@app.after_request
def add_no_store(resp):
    if request.path.startswith('/api/'):
        resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        resp.headers['Pragma'] = 'no-cache'
        resp.headers['Expires'] = '0'
    return resp
        
@app.get("/api/health")
def api_health():
    return {"ok": True, "time": datetime.utcnow().isoformat()}
# =========================
#  Run
# =========================
if __name__ == "__main__":
    # No hay warming de Sheets; BD ya está lista
    app.run(host="0.0.0.0", port=8000, debug=True)
